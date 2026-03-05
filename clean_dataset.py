"""
Pantun Dataset Cleaning & Consolidation Script
Processes 7 raw Excel files, normalizes themes to 11 standard categories,
merges with existing pantun_cleaned_v2.json, deduplicates, and exports.
"""

import openpyxl
import json
import os
import re
from collections import Counter

# ============================================================
# THEME NORMALIZATION MAPPING
# Maps raw theme names → 11 standardized theme categories
# Pantun with themes not in this mapping will be DISCARDED
# ============================================================

THEME_MAP = {
    # ------ PANTUN PERCINTAAN ------
    "Kasih": "PANTUN PERCINTAAN",
    "Nasib, Cinta, Rintihan, Rayuan": "PANTUN PERCINTAAN",
    "Pantun Kasih Sayang": "PANTUN PERCINTAAN",
    "Gelora Cinta": "PANTUN PERCINTAAN",
    "Menabur Cinta": "PANTUN PERCINTAAN",
    "Sejadah Cinta": "PANTUN PERCINTAAN",
    "Cinta Kertas Tisu": "PANTUN PERCINTAAN",
    "Cinta dan Pasrah": "PANTUN PERCINTAAN",
    "Gadis Julang Bersubang": "PANTUN PERCINTAAN",
    "Bujang Julung Berkeris": "PANTUN PERCINTAAN",
    "Selagi Ada Bulan dan Bintang": "PANTUN PERCINTAAN",
    "Keserasian": "PANTUN PERCINTAAN",
    "Intan dengan Kaca": "PANTUN PERCINTAAN",
    "Yang Cantik Adalah Hati": "PANTUN PERCINTAAN",

    # ------ PANTUN NASIHAT DAN PENDIDIKAN ------
    "Nasihat": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Pantun Nasihat": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Jangan Terpedaya": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Jangan Ada Benci": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Contohi Penyu": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Benih yang Baik": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Resmi Padi": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Sabar Itu Indah": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Kalau Kau Tahu": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Kata Mesti Dikota": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Buang yang Keruh": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Mengapa Jadi Begini": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Pentingkan Diri Sendiri": "PANTUN NASIHAT DAN PENDIDIKAN",
    "Tiada Lagi Watak Berlembut": "PANTUN NASIHAT DAN PENDIDIKAN",

    # ------ PANTUN AGAMA DAN KEPERCAYAAN ------
    "Agama": "PANTUN AGAMA DAN KEPERCAYAAN",
    "Pantun Berkait (Agama)": "PANTUN AGAMA DAN KEPERCAYAAN",
    "Demi Kalimah Allah": "PANTUN AGAMA DAN KEPERCAYAAN",
    "Bukan Zalim Suami Soleh": "PANTUN AGAMA DAN KEPERCAYAAN",
    "Ibu Solehah": "PANTUN AGAMA DAN KEPERCAYAAN",

    # ------ PANTUN BUDI ------
    "Budi": "PANTUN BUDI",
    "Pantun Budi dan Jasa": "PANTUN BUDI",
    "Pantun Setia Budi": "PANTUN BUDI",
    "Budi Dijunjung": "PANTUN BUDI",
    "Budi Bahasa Budaya Kita": "PANTUN BUDI",
    "Kacang Lupakan Kulit": "PANTUN BUDI",
    "Setia yang Waja": "PANTUN BUDI",
    "Sahabat Sejati": "PANTUN BUDI",
    "Tolong-menolong": "PANTUN BUDI",

    # ------ PANTUN JENAKA DAN PERMAINAN ------
    "Jenaka": "PANTUN JENAKA DAN PERMAINAN",
    "Pantun Permainan Hidup": "PANTUN JENAKA DAN PERMAINAN",
    "Kopi dan Pantun": "PANTUN JENAKA DAN PERMAINAN",
    "Sekuntum Pantun": "PANTUN JENAKA DAN PERMAINAN",

    # ------ PANTUN PERIBAHASA DAN PERBILANGAN ------
    "Peribahasa, Kata Hikmat, Puisi Berirama": "PANTUN PERIBAHASA DAN PERBILANGAN",
    "Pantun Peribahasa": "PANTUN PERIBAHASA DAN PERBILANGAN",
    "Ibarat Beruk Diberi Bunga": "PANTUN PERIBAHASA DAN PERBILANGAN",
    "Gading Bertuah": "PANTUN PERIBAHASA DAN PERBILANGAN",

    # ------ PANTUN KIAS DAN IBARAT ------
    "Sindiran, Kiasan, Kritis, Sinis": "PANTUN KIAS DAN IBARAT",
    "Pantun Kiasan": "PANTUN KIAS DAN IBARAT",
    "Lain di Mulut Lain di Hati": "PANTUN KIAS DAN IBARAT",
    "Berpura-pura": "PANTUN KIAS DAN IBARAT",
    "Musuh dalam Selimut": "PANTUN KIAS DAN IBARAT",
    "Duri dalam Daging": "PANTUN KIAS DAN IBARAT",
    "Sombong ibarat Tangga": "PANTUN KIAS DAN IBARAT",
    "Riak dan Bangga": "PANTUN KIAS DAN IBARAT",
    "Sifat Angkuh": "PANTUN KIAS DAN IBARAT",
    "Bukan seperti Lalang": "PANTUN KIAS DAN IBARAT",
    "Menegakkan Benang Basah": "PANTUN KIAS DAN IBARAT",
    "Kerbau Dicucuk Hidung": "PANTUN KIAS DAN IBARAT",
    "Berpatah Arang": "PANTUN KIAS DAN IBARAT",
    "Seperti Sehelai Tisu": "PANTUN KIAS DAN IBARAT",

    # ------ PANTUN ADAT DAN RESAM ------
    "Pantun Adat Bermasyarakat": "PANTUN ADAT DAN RESAM",
    "Pantun Pengacaraan Majlis": "PANTUN ADAT DAN RESAM",
    "Ukur Baju di Badan Sendiri": "PANTUN ADAT DAN RESAM",
    "Sebumbung Tidak Sehaluan": "PANTUN ADAT DAN RESAM",

    # ------ PANTUN KEPAHLAWANAN ------
    "Semangat, Taat, Harmonis": "PANTUN KEPAHLAWANAN",
    "Pantun Usaha dan Perjuangan": "PANTUN KEPAHLAWANAN",
    "Perjuangan Bangsa": "PANTUN KEPAHLAWANAN",
    "Memperkasa Bangsa": "PANTUN KEPAHLAWANAN",
    "Sifat Bangsaku": "PANTUN KEPAHLAWANAN",
    "Peluru Habis": "PANTUN KEPAHLAWANAN",
    "Utamakan Perpaduan": "PANTUN KEPAHLAWANAN",
    "Bersama Rakyat": "PANTUN KEPAHLAWANAN",
    "Negara Sejahtera Rakyat Bahagia": "PANTUN KEPAHLAWANAN",
    "Bahagia Menjadi Rakyat Malaysia": "PANTUN KEPAHLAWANAN",
    "Kenyalang yang Terbilang": "PANTUN KEPAHLAWANAN",

    # ------ PANTUN KEMBARA DAN PERANTAUAN ------
    "Mengejar Pelangi": "PANTUN KEMBARA DAN PERANTAUAN",
    "Pulang ke Pangkal Jalan": "PANTUN KEMBARA DAN PERANTAUAN",
    "Menggarap Harapan": "PANTUN KEMBARA DAN PERANTAUAN",

    # ------ PANTUN TEKA-TEKI ------
    "Umum, Teka-Teki, Humor": "PANTUN TEKA-TEKI",
    "Umum, Teka-teki, Jenaka": "PANTUN TEKA-TEKI",

    # ------ MIXED / MULTI-THEME (split into primary) ------
    # "Agama, Nasihat, Adat, Filsafat" is a mixed category from Pak Nazel books
    # We map to NASIHAT as the primary theme since nasihat is the dominant meaning
    "Agama, Nasihat, Adat, Filsafat": "PANTUN NASIHAT DAN PENDIDIKAN",

    # ------ PANTUN PERCINTAAN (family/emotion related) ------
    "Air Mata Ibu": "PANTUN PERCINTAAN",
    "Hati Ibu": "PANTUN PERCINTAAN",
    "Ibu Mithali": "PANTUN PERCINTAAN",
    "Suara Seorang Ibu": "PANTUN PERCINTAAN",
    "Setitis Susu Ibu": "PANTUN PERCINTAAN",
    "Anakku Anakmu": "PANTUN PERCINTAAN",

    # ------ PANTUN KIAS DAN IBARAT (emotion/metaphor) ------
    "Bara yang Membara": "PANTUN KIAS DAN IBARAT",
    "Dendam Masih Membara": "PANTUN KIAS DAN IBARAT",
    "Dendam Membara": "PANTUN KIAS DAN IBARAT",
    "Seperahu Dua Nakhoda": "PANTUN KIAS DAN IBARAT",

    # ------ PANTUN NASIHAT (general life advice titles) ------
    "Pantunku Suara Hatiku": "PANTUN NASIHAT DAN PENDIDIKAN",
}

# Themes that should be DISCARDED (cannot cleanly map)
DISCARD_THEMES = {
    "Mukadimah",  # Introduction/preamble, not a real theme
    "Hari ATM ke-80", "Hari ATM ke-81", "Hari ATM ke-82", "Hari ATM ke-83",
    "Hari ATM ke-84", "Hari ATM ke-85", "Hari ATM ke-86", "Hari ATM ke-87",
    "Hari ATM ke-88", "Hari ATM ke-89",
    "Azam Tahun Baharu",  # Very specific event
}

# The 11 standardized themes
VALID_THEMES = {
    "PANTUN ADAT DAN RESAM",
    "PANTUN AGAMA DAN KEPERCAYAAN",
    "PANTUN BUDI",
    "PANTUN JENAKA DAN PERMAINAN",
    "PANTUN KEMBARA DAN PERANTAUAN",
    "PANTUN KEPAHLAWANAN",
    "PANTUN KIAS DAN IBARAT",
    "PANTUN NASIHAT DAN PENDIDIKAN",
    "PANTUN PERCINTAAN",
    "PANTUN PERIBAHASA DAN PERBILANGAN",
    "PANTUN TEKA-TEKI",
}


def clean_pantun_text(text):
    """Clean and normalize pantun text."""
    if not text or not isinstance(text, str):
        return None
    
    # Strip whitespace
    text = text.strip()
    
    # Skip empty
    if not text:
        return None
    
    # Normalize multiple spaces
    text = re.sub(r'\s+', ' ', text)
    
    # Normalize various separators to semicolon with space
    # Some pantun use different separators between lines
    text = re.sub(r'\s*;\s*', '; ', text)
    text = re.sub(r'\s*,\s*', ', ', text)
    
    return text


def normalize_for_dedup(text):
    """Create a normalized version of text for deduplication."""
    if not text:
        return ""
    # Lowercase, remove all punctuation and extra spaces
    t = text.lower()
    t = re.sub(r'[^\w\s]', '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def read_excel_files(raw_dir):
    """Read all raw Excel files and extract pantun with themes."""
    raw_pantuns = []
    
    for filename in sorted(os.listdir(raw_dir)):
        if not filename.endswith('.xlsx'):
            continue
        
        filepath = os.path.join(raw_dir, filename)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Find column indices
        pantun_col = None
        theme_col = None
        
        for c in range(1, ws.max_column + 1):
            header = str(ws.cell(1, c).value or '').strip().upper()
            if 'PANTUN' in header:
                pantun_col = c
            if 'TEMA' in header or 'THEME' in header:
                theme_col = c
        
        if not pantun_col or not theme_col:
            print(f"  WARNING: Skipping {filename} - columns not found")
            continue
        
        file_count = 0
        file_discarded = 0
        
        for r in range(2, ws.max_row + 1):
            raw_theme = ws.cell(r, theme_col).value
            raw_pantun = ws.cell(r, pantun_col).value
            
            if not raw_theme or not raw_pantun:
                continue
            
            theme_str = str(raw_theme).strip()
            pantun_str = clean_pantun_text(str(raw_pantun))
            
            if not pantun_str:
                continue
            
            # Check if theme should be discarded
            if theme_str in DISCARD_THEMES:
                file_discarded += 1
                continue
            
            # Map to standardized theme
            mapped_theme = THEME_MAP.get(theme_str)
            
            if not mapped_theme:
                # Theme not in mapping - discard
                file_discarded += 1
                continue
            
            if mapped_theme not in VALID_THEMES:
                file_discarded += 1
                continue
            
            raw_pantuns.append({
                "tema": mapped_theme,
                "pantun": pantun_str,
                "sumber": filename.replace('.xlsx', '')
            })
            file_count += 1
        
        print(f"  {filename[:50]}: {file_count} kept, {file_discarded} discarded")
        wb.close()
    
    return raw_pantuns


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    raw_dir = os.path.join(base_dir, "raw")
    cleaned_json = os.path.join(base_dir, "pantun_cleaned_v2.json")
    output_json = os.path.join(base_dir, "pantun_dataset.json")
    
    print("=" * 60)
    print("PANTUN DATASET CLEANING & CONSOLIDATION")
    print("=" * 60)
    
    # Step 1: Read existing cleaned dataset
    print("\n[1/4] Loading existing cleaned dataset...")
    with open(cleaned_json, 'r', encoding='utf-8') as f:
        existing = json.load(f)
    print(f"  Loaded {len(existing)} pantun from pantun_cleaned_v2.json")
    
    # Step 2: Read raw Excel files
    print("\n[2/4] Processing raw Excel files...")
    raw_pantuns = read_excel_files(raw_dir)
    print(f"  Total from raw files: {len(raw_pantuns)}")
    
    # Step 3: Merge and deduplicate
    print("\n[3/4] Merging and deduplicating...")
    
    # Build dedup set from existing
    seen = set()
    final_dataset = []
    
    for item in existing:
        norm = normalize_for_dedup(item.get('pantun', ''))
        if norm and norm not in seen:
            seen.add(norm)
            final_dataset.append({
                "tema": item['tema'],
                "pantun": item['pantun']
            })
    
    existing_count = len(final_dataset)
    
    # Add raw pantuns (skip duplicates)
    added_from_raw = 0
    for item in raw_pantuns:
        norm = normalize_for_dedup(item['pantun'])
        if norm and norm not in seen:
            seen.add(norm)
            final_dataset.append({
                "tema": item['tema'],
                "pantun": item['pantun']
            })
            added_from_raw += 1
    
    print(f"  Existing (after dedup): {existing_count}")
    print(f"  Added from raw: {added_from_raw}")
    print(f"  Total final: {len(final_dataset)}")
    
    # Step 4: Export
    print("\n[4/4] Exporting final dataset...")
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(final_dataset, f, ensure_ascii=False, indent=2)
    print(f"  Saved to: {output_json}")
    
    # Statistics
    print("\n" + "=" * 60)
    print("FINAL DATASET STATISTICS")
    print("=" * 60)
    
    theme_counts = Counter(item['tema'] for item in final_dataset)
    for theme, count in theme_counts.most_common():
        bar = "█" * (count // 20)
        print(f"  {theme:45s} {count:5d}  {bar}")
    
    print(f"\n  {'TOTAL':45s} {len(final_dataset):5d}")
    print(f"  Unique themes: {len(theme_counts)}")
    
    # Check for any unmapped themes in raw
    print("\n" + "=" * 60)
    print("UNMAPPED THEMES CHECK")
    print("=" * 60)
    
    all_raw_themes = set()
    for f in os.listdir(raw_dir):
        if f.endswith('.xlsx'):
            wb = openpyxl.load_workbook(os.path.join(raw_dir, f))
            ws = wb.active
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(1, c).value or '').upper()
                if 'TEMA' in h:
                    for r in range(2, ws.max_row + 1):
                        v = ws.cell(r, c).value
                        if v:
                            all_raw_themes.add(str(v).strip())
            wb.close()
    
    unmapped = all_raw_themes - set(THEME_MAP.keys()) - DISCARD_THEMES
    if unmapped:
        print(f"  Found {len(unmapped)} unmapped themes (these were discarded):")
        for t in sorted(unmapped):
            print(f"    - {t}")
    else:
        print("  All themes accounted for (mapped or explicitly discarded).")


if __name__ == "__main__":
    main()
